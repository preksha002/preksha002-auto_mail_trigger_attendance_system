import face_recognition
import cv2
import openpyxl
from openpyxl import Workbook
import datetime
import os
import eel
from binascii import a2b_base64
import urllib.request
import yagmail

eel.init('web')

DIRECTORY_PATH = '/Users/Dell/OneDrive/Desktop/sms'                             # Put here the PROJECT DIRECTORY PATH
IMAGE_DIRECTORY_PATH = '/Users/Dell/OneDrive/Desktop/sms/assets/'                # Put here the ASSESTS FOLDER PATH


#-------------------------------------------------------------------------------------------------------------------------------------#
# MODULE - 1 (SAVING THE IMAGE DATA FROM JS -> PYTHON -> LOCAL STORAGE)
#-------------------------------------------------------------------------------------------------------------------------------------#

# Define Python function called in JS and expose it to eel, this method will save the screenshot image data with name image.jpg

@eel.expose
def get_image_data(data):
    response = urllib.request.urlopen(data)         # Convert the raw_image(DATA_URI to binary image to save it)
    with open(DIRECTORY_PATH + 'image.jpg', 'wb') as f:
        f.write(response.file.read())
    return "Python says thanks"




#-------------------------------------------------------------------------------------------------------------------------------------#
# MODULE - 2 (CREATING EXCEL SHEET AND CALCULATING DEPENDENCIES)
#-------------------------------------------------------------------------------------------------------------------------------------#

# Load everything i.e workbook, video capture (videocam)

# Load present date and time
now= datetime.datetime.now().date()


print(now)
today=now.day
print(today)
month=now.month
month_name = datetime.date(1900, month, 1).strftime('%B')           # Getting the month name from the month's integer


# Get a reference to webcam #0 (the default one)
video_capture = cv2.VideoCapture(0)
    
    
# Create a woorksheet only when the file does not exist else update the existing excel workbook
if (not os.path.isfile(DIRECTORY_PATH + str(now) + '.xlsx')):
    book = Workbook()
    sheet = book.active
else:
    book= openpyxl.load_workbook(str(now) + '.xlsx')
    sheet=book.active




#-------------------------------------------------------------------------------------------------------------------------------------#
# MODULE - 3 (TRAINING AND LOADING KNOWN FACE ENCODINGS)
#-------------------------------------------------------------------------------------------------------------------------------------#

@eel.expose
def save_student_data(roll_number,name,clas):
    os.rename(DIRECTORY_PATH + 'image.jpg',IMAGE_DIRECTORY_PATH + str(name) +"_"+str(roll_number)+"-"+str(clas) + '.jpg')    # This will change the image.jpg to the roll number inserted in field.jpg

        

# Create arrays of known face encodings and their names
known_face_encodings = []
known_face_names = []

print('Following images have been trained : ')
for filename in os.listdir(IMAGE_DIRECTORY_PATH):
    print(str(filename))
    if filename.endswith(".jpg"):
        image = face_recognition.load_image_file(IMAGE_DIRECTORY_PATH + str(filename))                 # Load Image
        image_face_encoding = face_recognition.face_encodings(image)[0]         # Find the face encoding
        known_face_encodings.append(image_face_encoding)                        # Append the face encoding to the known faces
        known_face_names.append(filename)                                       # Append the name to the known face names
    


@eel.expose
def sendautomaticemail():
    try:
        print("Sending Email .. ")
    #initializing the server connection
        yag = yagmail.SMTP(user='sv4694438@gmail.com', password='Shally002')
    #sending the email
        yag.send(to=['vermavivekkumar001@gmail.com','prekshawork002@gmail.com','vivekhbti428@gmail.com'] , subject='Attendance Sheet :'+str(now), contents='Dear All , \n Please find the file attached for today attendance '+str(now)+ ':-', attachments='C:/Users/Dell/OneDrive/Desktop/sms/'+str(now)+'.xlsx')
        print("Email sent successfully")
    except:
        print("Error, email was not sent")
   
#-------------------------------------------------------------------------------------------------------------------------------------#
# MODULE - 4 (STARTING THE WEB CAM STREAM FOR TAKING ATTENDANCE, PRESS CTRL + C TO STOP )
#-------------------------------------------------------------------------------------------------------------------------------------#

# This method will be used starting the video stream for taking the attendance of all the already known faces

@eel.expose
def take_attendance():
    # Initialize some variables
    face_locations = []
    face_encodings = []
    face_names = []
    process_this_frame = True
    sheet.cell(row=int(1), column=int(1)).value = "Roll No"
    sheet.cell(row=int(1), column=int(2)).value = "Student Name"
    sheet.cell(row=int(1), column=int(3)).value = "Class"
    sheet.cell(row=int(1), column=int(4)).value = "Attendance Status"
    
    while True:
        
        # Grab a single frame of video
        ret, frame = video_capture.read()
        
        # Resize frame of video to 1/4 size for faster face recognition processing
        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        
        # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
        rgb_small_frame = small_frame[:, :, ::-1]
        
        # Only process every other frame of video to save time
        if process_this_frame:
            # Find all the faces and face encodings in the current frame of video
            face_locations = face_recognition.face_locations(rgb_small_frame)
            face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
        
        face_names = []
        name = ""
        student_name = ""
        for face_encoding in face_encodings:
            # See if the face is a match for the known face(s)
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            name = "Unknown"
        
            # If a match was found in known_face_encodings, just use the first one.
            if True in matches:
                first_match_index = matches.index(True)
                name = known_face_names[first_match_index]
                #print(name)
                student_name = name[0:name.find('_')]
                #print(student_name) #student name
                xyz = name[name.find('_')+1:name.find('-')]
                clss = name[name.find('-')+1:name.find('.')]
                #print(clss)
                #print(xyz) #roll no
                # Assign attendance
                
                if int(xyz) in range(1,61):
                    sheet.cell(row=int(xyz)+1, column=int(4)).value = "P"
                    sheet.cell(row=int(xyz)+1, column=int(1)).value = xyz
                    sheet.cell(row=int(xyz)+1, column=int(2)).value = student_name
                    sheet.cell(row=int(xyz)+1, column=int(3)).value = clss
            else:
                sheet.cell(row=int(xyz)+1, column=int(4)).value = "A"
                sheet.cell(row=int(xyz)+1, column=int(1)).value = xyz
                sheet.cell(row=int(xyz)+1, column=int(2)).value = student_name
                sheet.cell(row=int(xyz)+1, column=int(3)).value = clss
        
        face_names.append(student_name)
        
        process_this_frame = not process_this_frame
        
        top, right, left, bottom = 1,1,1,1      # Default values since the variables not defined were giving error
        # Display the results
        for (top, right, bottom, left), name in zip(face_locations, face_names):
            # Scale back up face locations since the frame we detected in was scaled to 1/4 size
            top *= 4
            right *= 4
            bottom *= 4
            left *= 4
        
        # Draw a box around the face
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
        
            # Draw a label with a name below the face
        cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
        font = cv2.FONT_HERSHEY_DUPLEX
        cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)
        
        # Display the resulting image
        cv2.imshow('Video', frame)
            
        # Save Woorksheet as present month
        book.save(str(now)+'.xlsx')
        
        # Hit 'q' on the keyboard to quit!
        if cv2.waitKey(1) & 0xFF == ord('q'):
            sendautomaticemail()
            break
        
    # Release handle to the webcam
    video_capture.release()
    cv2.destroyAllWindows()




#-------------------------------------------------------------------------------------------------------------------------------------#
# MODULE - 5 (EEL STARTING THE SERVER)
#-------------------------------------------------------------------------------------------------------------------------------------#

# To start the web server of the eel 

eel.start('index.html', mode='chrome-app', port=8080, cmdline_args=['--start-fullscreen', '--browser-startup-dialog'])